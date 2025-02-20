import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image
import base64
import os
from pathlib import Path
import openpyxl
from datetime import datetime

class ImageConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertidor de Imágenes a Base64")
        self.root.geometry("800x600")
        
        # Variables
        self.folder_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.image_list = []
        self.max_file_size = 5 * 1024 * 1024  # 5MB por defecto
        
        # Crear la interfaz
        self.create_widgets()
        
    def create_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Selector de carpeta de entrada
        ttk.Label(main_frame, text="Carpeta de imágenes:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(main_frame, textvariable=self.folder_path, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(main_frame, text="Buscar", command=self.select_input_folder).grid(row=0, column=2)
        
        # Selector de carpeta de salida
        ttk.Label(main_frame, text="Guardar Excel en:").grid(row=1, column=0, sticky=tk.W)
        ttk.Entry(main_frame, textvariable=self.output_path, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(main_frame, text="Buscar", command=self.select_output_folder).grid(row=1, column=2)
        
        # Lista de archivos
        ttk.Label(main_frame, text="Imágenes encontradas:").grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(10,0))
        
        # Frame para la lista con scroll
        list_frame = ttk.Frame(main_frame)
        list_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.file_listbox = tk.Listbox(list_frame, width=70, height=15)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.file_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Barra de progreso
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, length=300, mode='determinate', variable=self.progress_var)
        self.progress_bar.grid(row=4, column=0, columnspan=3, pady=10, sticky=(tk.W, tk.E))
        
        # Etiqueta de estado
        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.grid(row=5, column=0, columnspan=3, sticky=tk.W)
        
        # Botón de conversión
        ttk.Button(main_frame, text="Convertir Imágenes", command=self.convert_images).grid(row=6, column=0, columnspan=3, pady=10)
        
        # Configurar el grid
        main_frame.columnconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
    def select_input_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_path.set(folder)
            self.update_file_list()
            
    def select_output_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_path.set(folder)
            
    def update_file_list(self):
        self.file_listbox.delete(0, tk.END)
        self.image_list = []
        
        if not self.folder_path.get():
            return
            
        valid_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp')
        
        for file in os.listdir(self.folder_path.get()):
            if file.lower().endswith(valid_extensions):
                full_path = os.path.join(self.folder_path.get(), file)
                if os.path.isfile(full_path):
                    self.image_list.append(full_path)
                    self.file_listbox.insert(tk.END, file)
                    
        self.status_label.config(text=f"Encontradas {len(self.image_list)} imágenes")
        
    def convert_images(self):
        if not self.image_list:
            messagebox.showwarning("Advertencia", "No hay imágenes para convertir")
            return
            
        if not self.output_path.get():
            messagebox.showwarning("Advertencia", "Por favor seleccione una carpeta de destino")
            return
            
        # Crear nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Imágenes"
        
        # Agregar encabezados
        ws['A1'] = "Nombre de Archivo"
        ws['B1'] = "Imagen Base64"
        
        total_files = len(self.image_list)
        self.progress_var.set(0)
        
        for i, image_path in enumerate(self.image_list):
            try:
                # Actualizar progreso
                self.progress_var.set((i + 1) / total_files * 100)
                self.status_label.config(text=f"Procesando: {os.path.basename(image_path)}")
                self.root.update()
                
                # Verificar tamaño del archivo
                file_size = os.path.getsize(image_path)
                if file_size > self.max_file_size:
                    if not messagebox.askyesno("Advertencia", 
                        f"El archivo {os.path.basename(image_path)} excede el tamaño recomendado "
                        f"({file_size/1024/1024:.2f}MB). ¿Desea continuar con este archivo?"):
                        continue
                
                # Leer y convertir imagen
                with open(image_path, 'rb') as image_file:
                    encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                
                # Agregar a Excel
                ws.append([os.path.basename(image_path), encoded_string])
                
            except Exception as e:
                messagebox.showerror("Error", f"Error procesando {os.path.basename(image_path)}: {str(e)}")
                
        # Guardar Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(self.output_path.get(), f"Imagenes_Base64_{timestamp}.xlsx")
        
        try:
            wb.save(excel_path)
            self.status_label.config(text="Proceso completado exitosamente")
            messagebox.showinfo("Éxito", f"Archivo Excel guardado en:\n{excel_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Error guardando el archivo Excel: {str(e)}")
            
        self.progress_var.set(100)

if __name__ == "__main__":
    root = tk.Tk()
    app = ImageConverterApp(root)
    root.mainloop()